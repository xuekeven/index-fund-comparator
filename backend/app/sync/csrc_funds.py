import argparse
import html
import io
import json
import logging
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any
from urllib.parse import quote, unquote, urljoin
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session

from app.database import get_session_factory
from app.database_models import (
    FundProduct,
    FundScale,
    FundShareClass,
    NavDaily,
    SalesLimitHistory,
)
from app.sync_history import run_tracked_sync
from app.sync.company_official_nav import OfficialCompanyNavFetcher
from app.sync.eid_disclosures import (
    DisclosureDocument,
    disclosure_documents as find_disclosure_documents,
    fetch_disclosure_text,
    parse_fee_rates,
    sync_fee_history,
)
from app.sync.sse_details import (
    calculate_return_metrics,
    load_nav_records as load_return_nav_records,
    sync_return_metrics,
)
from app.sync.sse_funds import ensure_index_master_data


CSRC_INDEX_PAGE_URL = "https://www.csrc.gov.cn/csrc/c101900/c1029655/content.shtml"
EID_BASE_URL = "http://eid.csrc.gov.cn/fund"
EID_VALIDATE_URL = f"{EID_BASE_URL}/disclose/validate_fund.do"
EID_DETAIL_URL = f"{EID_BASE_URL}/disclose/fund_detail.do"
EID_ADVANCED_SEARCH_URL = f"{EID_BASE_URL}/disclose/advanced_search_report.do"
EID_SEARCH_META_URL = f"{EID_BASE_URL}/disclose/publicDailyReportSearchData.json"
EID_NAV_URL = f"{EID_BASE_URL}/disclose/getPublicFundJZInfoMore.do"
ASIA_SHANGHAI = ZoneInfo("Asia/Shanghai")
PRODUCT_SUMMARY_LABEL = "基金产品资料概要"
QUARTERLY_REPORT_LABEL = "季度报告"
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class TargetIndex:
    family_id: str
    definition_id: str
    benchmark_description: str
    investment_scopes: tuple[str, ...]


@dataclass(frozen=True)
class ProductCandidate:
    code: str
    name: str
    inception_date: date | None
    target: TargetIndex
    product_structure: str


@dataclass(frozen=True)
class FundDetail:
    fund_id: int
    code: str
    short_name: str
    full_name: str
    manager: str
    inception_date: date | None


@dataclass(frozen=True)
class NavRecord:
    code: str
    display_name: str
    share_class: str | None
    currency: str
    currency_form: str | None
    nav_date: date
    unit_nav: Decimal
    accumulated_nav: Decimal | None
    source_url: str = EID_NAV_URL


@dataclass(frozen=True)
class SummaryShare:
    code: str
    display_name: str
    share_class: str | None
    currency: str
    currency_form: str | None


@dataclass(frozen=True)
class ProductSummary:
    shares: tuple[SummaryShare, ...]
    rates: dict[str, Decimal]
    benchmark_description: str | None
    rates_by_code: dict[str, dict[str, Decimal]]

    @property
    def share_codes(self) -> tuple[str, ...]:
        return tuple(share.code for share in self.shares)


@dataclass(frozen=True)
class SubscriptionState:
    code: str
    status: str
    limit_amount: Decimal | None
    currency: str
    effective_date: date
    source_url: str
    channel: str = "全部渠道"
    quality_status: str = "verified"



@dataclass(frozen=True)
class ScaleRecord:
    share_code: str
    report_date: date
    amount_cny: Decimal


@dataclass(frozen=True)
class SyncStats:
    products: int
    shares: int
    nav_rows: int
    fee_shares: int
    scales: int
    subscription_states: int
    failures: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()
    return_metrics: int = 0


@dataclass(frozen=True)
class CatalogSyncStats:
    products: int
    shares: int
    snapshot_date: date
    failures: tuple[str, ...] = ()
    fee_shares: int = 0
    scales: int = 0
    warnings: tuple[str, ...] = ()


CSI_500 = TargetIndex(
    "csi-500",
    "csi-500-price-cny",
    "中证500价格指数（依据基金名称归类，精确合同口径待核验）",
    ("境内",),
)
SP_500 = TargetIndex(
    "sp-500",
    "sp-500-price-cny",
    "标普500指数（人民币折算及收益口径待基金合同核验）",
    ("QDII",),
)
NASDAQ_100 = TargetIndex(
    "nasdaq-100",
    "nasdaq-100-price-cny",
    "纳斯达克100指数（人民币折算及收益口径待基金合同核验）",
    ("QDII",),
)

TARGET_BY_DEFINITION = {
    target.definition_id: target
    for target in (CSI_500, SP_500, NASDAQ_100)
}

CSI_EXCLUDED_TOKENS = (
    "增强", "量化", "优选", "智选", "行业中性", "低波", "等权", "质量",
    "成长", "价值", "基本面", "ESG", "信息", "自由现金流", "策略",
)
US_EXCLUDED_TOKENS = ("等权", "低波", "质量", "增强", "策略", "科技", "生物")


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", "", value).replace("（", "(").replace("）", ")").upper()


def is_exchange_traded_summary(text: str) -> bool:
    compact = re.sub(r"\s+", "", text)
    listing = re.search(
        r"上市交易所及上市日期(.+?)(?:基金类型|基金类别)", compact
    )
    if listing:
        value = re.sub(
            r"^[（(]若有[）)]",
            "",
            listing.group(1),
        ).strip("-—－/、")
        if value and value not in {"无", "不适用", "未上市", "暂未上市"}:
            return True
    operation = re.search(
        r"运作方式(.+?)(?:开放频率|基金经理|基金合同存续期)", compact
    )
    return bool(
        operation
        and any(token in operation.group(1) for token in ("交易型开放式", "上市开放式"))
    )


def classify_product(name: str) -> tuple[TargetIndex, str] | None:
    normalized = normalize_name(name)
    if "LOF" in normalized or ("ETF" in normalized and "联接" not in normalized):
        return None
    structure = (
        "ETF联接基金"
        if "ETF" in normalized and "联接" in normalized
        else "普通开放式指数基金"
    )
    if "中证500" in normalized:
        if any(token.upper() in normalized for token in CSI_EXCLUDED_TOKENS):
            return None
        return CSI_500, structure
    if "标普500" in normalized:
        if any(token.upper() in normalized for token in US_EXCLUDED_TOKENS):
            return None
        return SP_500, structure
    if any(token in normalized for token in ("纳斯达克100", "纳指100")):
        if any(token.upper() in normalized for token in US_EXCLUDED_TOKENS):
            return None
        return NASDAQ_100, structure
    return None


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def snapshot_date_from_url(url: str) -> date:
    match = re.search(r"(20\d{6})", unquote(url))
    if match is None:
        raise RuntimeError("CSRC product index URL did not contain a snapshot date")
    return datetime.strptime(match.group(1), "%Y%m%d").date()


def parse_product_index(
    content: bytes, *, fallback_snapshot_date: date | None = None
) -> tuple[list[ProductCandidate], date]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    title = str(sheet.cell(1, 1).value or "")
    snapshot_match = re.search(r"截至\s*(\d{8})", title)
    if snapshot_match is None and fallback_snapshot_date is None:
        raise RuntimeError("CSRC product index did not contain a snapshot date")
    snapshot_date = (
        datetime.strptime(snapshot_match.group(1), "%Y%m%d").date()
        if snapshot_match
        else fallback_snapshot_date
    )
    assert snapshot_date is not None

    candidates: list[ProductCandidate] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        raw_code, raw_name = row[1], row[2]
        if raw_code in (None, "") or raw_name in (None, ""):
            continue
        code = str(raw_code).strip()
        if code.isdigit():
            code = code.zfill(6)
        if not re.fullmatch(r"\d{6}", code):
            continue
        name = str(raw_name).strip()
        classified = classify_product(name)
        if classified is None:
            continue
        target, structure = classified
        candidates.append(
            ProductCandidate(
                code, name, _as_date(row[3]), target, structure
            )
        )
    return candidates, snapshot_date


def extract_index_download_url(page_html: str) -> str:
    links = re.findall(
        r'href=["\']([^"\']+\.xlsx(?:\?[^"\']*)?)["\']', page_html, re.I
    )
    if not links:
        raise RuntimeError("CSRC product index page did not contain an XLSX download")
    return urljoin(CSRC_INDEX_PAGE_URL, html.unescape(links[-1]))


def fetch_product_index(client: httpx.Client) -> tuple[bytes, str]:
    page = client.get(CSRC_INDEX_PAGE_URL)
    page.raise_for_status()
    download_url = extract_index_download_url(page.text)
    response = client.get(download_url)
    response.raise_for_status()
    return response.content, download_url


def _clean_html_cell(value: str) -> str:
    text = html.unescape(re.sub(r"<[^>]+>", "", value))
    return re.sub(r"\s+", " ", text).strip()


def _detail_field(page_html: str, label: str) -> str:
    pattern = rf">\s*{re.escape(label)}\s*</td>\s*<td[^>]*>(.*?)</td>"
    match = re.search(pattern, page_html, re.I | re.S)
    return _clean_html_cell(match.group(1)) if match else ""


def parse_fund_detail(fund_id: int, page_html: str) -> FundDetail:
    title_match = re.search(
        r'class=["\']title_tu["\'][^>]*>(.*?)</td>', page_html, re.I | re.S
    )
    heading = _clean_html_cell(title_match.group(1)) if title_match else ""
    heading_match = re.match(r"(.+?)\((\d{6})\)$", heading)
    code = _detail_field(page_html, "基金代码")
    short_name = heading_match.group(1) if heading_match else ""
    if not code and heading_match:
        code = heading_match.group(2)
    full_name = _detail_field(page_html, "基金名称") or short_name
    manager = _detail_field(page_html, "基金管理人")
    inception_date = _as_date(_detail_field(page_html, "基金合同生效日期"))
    if not all((code, full_name, manager)):
        raise RuntimeError(f"EID detail page was incomplete for fundId={fund_id}")
    return FundDetail(
        fund_id, code, short_name or full_name, full_name, manager, inception_date
    )


def validate_fund_code(client: httpx.Client, code: str) -> int:
    response = client.post(EID_VALIDATE_URL, data={"cFundCode": code})
    response.raise_for_status()
    payload = response.json()
    if not payload.get("isSuccess") or not payload.get("fundId"):
        raise RuntimeError(f"EID did not recognize fund code {code}")
    return int(payload["fundId"])


def fetch_fund_detail(client: httpx.Client, fund_id: int) -> FundDetail:
    response = client.get(EID_DETAIL_URL, params={"fundId": fund_id})
    response.raise_for_status()
    return parse_fund_detail(fund_id, response.text)


def fetch_fund_detail_document(
    client: httpx.Client, fund_id: int
) -> tuple[FundDetail, str]:
    response = client.get(EID_DETAIL_URL, params={"fundId": fund_id})
    response.raise_for_status()
    return parse_fund_detail(fund_id, response.text), response.text


def fund_detail_url(fund_id: int) -> str:
    return f"{EID_DETAIL_URL}?fundId={fund_id}"


def disclosure_documents(page_html: str, title_token: str) -> list[DisclosureDocument]:
    return find_disclosure_documents(
        page_html,
        title_token,
        base_url=EID_DETAIL_URL,
    )


def product_summary_documents(
    client: httpx.Client,
    detail_html: str,
    fund_code: str,
    end_date: date,
) -> list[DisclosureDocument]:
    """Return current summaries, including files omitted by the detail-page preview."""
    query = [
        {"name": "sEcho", "value": 1},
        {"name": "iDisplayStart", "value": 0},
        {"name": "iDisplayLength", "value": 100},
        {"name": "fundType", "value": ""},
        {"name": "reportType", "value": "FA"},
        {"name": "reportYear", "value": ""},
        {"name": "fundCompanyShortName", "value": ""},
        {"name": "fundCode", "value": fund_code},
        {"name": "fundShortName", "value": ""},
        {
            "name": "startUploadDate",
            "value": (end_date - timedelta(days=400)).isoformat(),
        },
        {"name": "endUploadDate", "value": end_date.isoformat()},
    ]
    response = client.get(
        EID_ADVANCED_SEARCH_URL,
        params={"aoData": json.dumps(query, ensure_ascii=False)},
    )
    response.raise_for_status()

    documents: list[DisclosureDocument] = []
    for row in response.json().get("aaData", []):
        title = str(row.get("reportName") or "").strip()
        instance_id = row.get("uploadInfoId")
        if PRODUCT_SUMMARY_LABEL not in title or instance_id is None:
            continue
        documents.append(
            DisclosureDocument(
                title=title,
                url=(
                    f"{EID_BASE_URL}/disclose/instance_show_pdf_id.do"
                    f"?instanceid={int(instance_id)}"
                ),
            )
        )

    documents.extend(disclosure_documents(detail_html, PRODUCT_SUMMARY_LABEL))
    return list({document.url: document for document in documents}.values())


def _code_chunks(value: str) -> tuple[str, ...]:
    return tuple(value[index:index + 6] for index in range(0, len(value), 6))


_ANNOUNCEMENT_CODE_LABEL_PATTERN = (
    r"(?:下属分级基金的交易代码|下属基金份额的交易代码|"
    r"下属基金的交易代码|下属基金交易代码|"
    r"涉及基金份额类别的交易代码|各基金份额类别(?:的)?交易代码)"
)


def _announcement_table_codes(compact: str) -> tuple[str, ...]:
    """Read share codes from PDF tables, including currency-annotated cells."""
    label_match = re.search(_ANNOUNCEMENT_CODE_LABEL_PATTERN, compact)
    if label_match:
        tail = compact[label_match.end():label_match.end() + 320]
        boundary_match = re.search(
            r"(?:该分级基金是否|该基金份额是否|该基金份额类别是否|"
            r"各基金份额类别是否|下属分级基金是否|限制申购金额|"
            r"金额单位|2[、.]其他)",
            tail,
        )
        code_block = tail[:boundary_match.start()] if boundary_match else tail
        codes: list[str] = []
        for digit_run in re.findall(r"(?<!\d)\d{6,}(?!\d)", code_block):
            if len(digit_run) % 6 == 0:
                codes.extend(_code_chunks(digit_run))
        if codes:
            return tuple(dict.fromkeys(codes))

    # Some PDF tables are extracted column-first, leaving the label after
    # the concatenated codes: ``018966...021773额的交易代码``.
    reverse_match = re.search(r"((?:\d{6})+)额的交易代码", compact)
    return _code_chunks(reverse_match.group(1)) if reverse_match else ()


def _summary_share_name(compact: str, code: str) -> str:
    patterns = (
        rf"下属基金简称(.+?)下属基金(?:交易)?代码{code}",
        rf"基金简称(?:[A-Z])?(.+?)基金代码(?:[A-Z])?{code}",
    )
    for pattern in patterns:
        matches = re.findall(pattern, compact)
        if matches:
            return matches[-1].strip("：:，,；;")
    return code


def _summary_shares(compact: str) -> tuple[SummaryShare, ...]:
    subordinate_groups = re.findall(
        r"下属基金简称(.+?)下属基金(?:交易)?代码((?:\d{6})+)",
        compact,
    )
    share_names: list[tuple[str, str]] = []
    for name_block, raw_codes in subordinate_groups:
        codes = _code_chunks(raw_codes)
        class_tokens = re.findall(r"[A-Z]", name_block.upper())
        for index, code in enumerate(codes):
            if len(codes) == 1:
                display_name = name_block.strip("：:，,；;")
            elif len(class_tokens) == len(codes):
                display_name = class_tokens[index]
            else:
                display_name = _summary_share_name(compact, code)
            share_names.append((code, display_name))

    if not share_names:
        for share_class, display_name, code in re.findall(
            r"基金简称([A-Z])(.{1,160}?)"
            r"基金代码\1(?:人民币|美元现汇|美元现钞)?(\d{6})",
            compact[:2000],
        ):
            normalized_name = display_name.strip("：:，,；;")
            if share_class not in normalized_name.upper():
                normalized_name = f"{normalized_name}{share_class}"
            share_names.append((code, normalized_name))

    if not share_names:
        subordinate = re.search(
            r"下属基金(?:交易)?代码((?:\d{6})+)", compact
        )
        share_code = re.search(r"份额代码(\d{6})", compact)
        primary = re.search(r"(?<!下属)基金代码(\d{6})", compact)
        if subordinate:
            share_names.extend(
                (code, _summary_share_name(compact, code))
                for code in _code_chunks(subordinate.group(1))
            )
        elif share_code:
            code = share_code.group(1)
            share_names.append((code, _summary_share_name(compact, code)))
        elif primary:
            code = primary.group(1)
            share_names.append((code, _summary_share_name(compact, code)))
        else:
            raise RuntimeError("Fund product summary did not contain a share code")

    return tuple(
        SummaryShare(
            code=code,
            display_name=display_name,
            share_class=share_identity(display_name)[0],
            currency=share_identity(display_name)[1],
            currency_form=share_identity(display_name)[2],
        )
        for code, display_name in dict(share_names).items()
    )


def _summary_rates_by_code(
    compact: str,
    shares: tuple[SummaryShare, ...],
    rates: dict[str, Decimal],
) -> dict[str, dict[str, Decimal]]:
    common_rates = {
        fee_type: rate
        for fee_type, rate in rates.items()
        if fee_type in {"management", "custody"}
    }
    rates_by_code = {share.code: dict(common_rates) for share in shares}

    zero_sales_classes: set[str] = set()
    for match in re.finditer(
        r"本基金(.{0,50}?)基金份额不收取销售服务费",
        compact,
    ):
        zero_sales_classes.update(re.findall(r"([A-Z])类", match.group(1)))
    for share in shares:
        if share.share_class in zero_sales_classes:
            rates_by_code[share.code]["sales_service"] = Decimal("0")

    sales_row = re.search(
        r"销售服务费(?P<label>.{0,180}?)(?P<rate>[0-9]+(?:\.[0-9]+)?)%",
        compact,
    )
    if sales_row:
        sales_rate = Decimal(sales_row.group("rate"))
        label = sales_row.group("label")
        explicit_shares = [
            share
            for share in shares
            if re.sub(r"\s+", "", share.display_name) in label
        ]
        if explicit_shares:
            for share in explicit_shares:
                rates_by_code[share.code]["sales_service"] = sales_rate
        else:
            for share in shares:
                if share.share_class not in zero_sales_classes:
                    rates_by_code[share.code]["sales_service"] = sales_rate
    elif len(shares) == 1:
        rates_by_code[shares[0].code]["sales_service"] = rates.get(
            "sales_service", Decimal("0")
        )

    found_comprehensive = False
    for share in shares:
        display_name = re.sub(r"\s+", "", share.display_name)
        match = re.search(
            rf"{re.escape(display_name)}基金运作综合费率"
            r"(?:[（(]年化[）)])?.{0,80}?([0-9]+(?:\.[0-9]+)?)%",
            compact,
        )
        if match:
            rates_by_code[share.code]["comprehensive_operating"] = Decimal(
                match.group(1)
            )
            found_comprehensive = True
    if not found_comprehensive and len(shares) == 1:
        comprehensive = rates.get("comprehensive_operating")
        if comprehensive is not None:
            rates_by_code[shares[0].code]["comprehensive_operating"] = comprehensive

    return rates_by_code


def parse_product_summary(text: str) -> ProductSummary:
    compact = re.sub(r"\s+", "", text)
    shares = _summary_shares(compact)
    rates = parse_fee_rates(text, include_sales_service=True)
    rates_by_code = _summary_rates_by_code(compact, shares, rates)

    benchmark_match = re.search(
        r"业绩比较基准(.+?)(?:风险收益特征|基金风险等级|\(二\)投资组合)", compact
    )
    benchmark = benchmark_match.group(1).strip("。；;") if benchmark_match else None
    return ProductSummary(shares, rates, benchmark or None, rates_by_code)


def _announcement_date(compact: str) -> date:
    date_pattern = r"(20\d{2})[年/-](\d{1,2})[月/-](\d{1,2})日?"
    effective_match = re.search(
        rf"(?:暂停(?:[（(]?大额[）)]?)?申购起始日|"
        rf"恢复(?:[（(]?大额[）)]?)?申购(?:起始)?日|自)"
        rf".{{0,20}}?{date_pattern}",
        compact,
    )
    matches = [effective_match.groups()] if effective_match else re.findall(
        date_pattern, compact
    )
    if not matches:
        raise RuntimeError("Subscription announcement did not contain a date")
    year, month, day = matches[0]
    return date(int(year), int(month), int(day))


def _announcement_share_codes(
    compact: str,
    title: str,
    shares: dict[str, SummaryShare],
) -> list[str]:
    normalized_title = re.sub(r"\s+", "", title)
    currency_scope: str | None = None
    if re.search(r"人民币(?:份额)?.{0,8}?(?:申购|定投)", normalized_title):
        currency_scope = "人民币"
    elif re.search(
        r"美元(?:现汇|现钞)?(?:份额)?.{0,8}?(?:申购|定投)",
        normalized_title,
    ):
        currency_scope = "美元"

    def share_is_in_scope(share: SummaryShare) -> bool:
        return currency_scope is None or share.currency == currency_scope

    normalized_upper_title = normalized_title.upper()
    classes = set(re.findall(r"([A-Z])类", normalized_upper_title))
    class_match = re.search(
        r"((?:[A-Z][、和及]?)+)(?:两|三|四)?类(?:基金)?份额",
        normalized_upper_title,
    )
    if class_match:
        classes.update(re.findall(r"[A-Z]", class_match.group(1)))
    if classes:
        return [
            code
            for code, share in shares.items()
            if share.share_class in classes and share_is_in_scope(share)
        ]

    codes = list(_announcement_table_codes(compact))
    codes = [
        code for code in codes
        if code in shares and share_is_in_scope(shares[code])
    ]
    if codes:
        flag_match = re.search(
            r"(?:该分级基金是否|该基金份额是否|该基金份额类别是否|"
            r"各基金份额类别是否|下属分级基金是否)"
            r"[^是]{0,160}([是否-]+)(?:2、|2\.|其他|金额单位|下属)",
            compact,
        )
        if flag_match and len(flag_match.group(1)) == len(codes):
            codes = [
                code
                for code, flag in zip(codes, flag_match.group(1), strict=True)
                if flag == "是"
            ]
        return codes

    class_prefix = title.split("类份额", 1)[0][-12:]
    classes = set(re.findall(r"([A-Z])(?:、|和|及|两|三|四|类)", class_prefix.upper()))
    if classes:
        return [
            code
            for code, share in shares.items()
            if share.share_class in classes and share_is_in_scope(share)
        ]
    return [code for code, share in shares.items() if share_is_in_scope(share)]


_LIMIT_UNIT_PATTERN = (
    r"(?:元人民币|人民币(?:亿元|万元|元)?|(?:亿|万)?美元|亿元|万元|元)"
)


def _limit_currency(unit: str) -> str:
    return "美元" if "美元" in unit else "人民币"


def _limit_amount(raw_amount: str, unit: str) -> Decimal:
    amount = Decimal(raw_amount.replace(",", ""))
    if "亿" in unit:
        return amount * Decimal("100000000")
    if "万" in unit:
        return amount * Decimal("10000")
    return amount


def _announcement_limits(
    text: str,
    compact: str,
    shares: dict[str, SummaryShare],
) -> dict[str, Decimal]:
    limits: dict[str, Decimal] = {}
    ordered_codes = [
        code
        for code in _announcement_table_codes(compact)
        if code in shares
    ]

    # Keep PDF row/column whitespace long enough to distinguish values such
    # as 100 100. Compacting first turns that row into 100100.
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        # PDF table headers commonly wrap ``金额`` and ``单位`` onto
        # separate lines. Locate the line that closes the unit marker, then
        # inspect only its tail and the immediately following value row. A
        # wider scan can mistake the next table's six-digit fund codes for
        # per-share amounts.
        header_end: int | None = None
        header_unit: str | None = None
        header_block = ""
        for candidate_end in range(index, min(index + 4, len(lines))):
            header_block = re.sub(
                r"\s+", "", "".join(lines[index : candidate_end + 1])
            )
            unit_match = re.search(
                rf"单位[：:]?(?P<unit>{_LIMIT_UNIT_PATTERN})[）)]",
                header_block,
            )
            if (
                "金额" in header_block
                and unit_match
                and (
                    "申购" in header_block
                    or "购金额" in header_block
                    or "限" in header_block
                )
            ):
                header_end = candidate_end
                header_unit = unit_match.group("unit")
                break
        if header_end is None or header_unit is None:
            continue
        value_lines: list[str] = []
        inline_value = re.split(r"[）)]", lines[header_end])[-1]
        if inline_value.strip():
            value_lines.append(inline_value)
        if header_end + 1 < len(lines):
            value_lines.append(lines[header_end + 1])
        for value_line in value_lines:
            values = re.findall(
                r"(?<!\d)(\d[\d,]*(?:\.\d+)?)(?!\d)", value_line
            )
            if len(values) != len(ordered_codes):
                continue
            for code, raw_amount in zip(ordered_codes, values, strict=True):
                limits[code] = _limit_amount(raw_amount, header_unit)
            break
        if limits:
            break
    table_match = re.search(
        r"(?:下属分级基金的限制申购金额|下属基金份额的限制金额|限制申购金额)"
        r"(?:（[^）]+）)?(.+?)(?:下属.*?限制(?:转换|定期)|2[.、]其他)",
        compact,
    )
    if table_match:
        table_values = re.findall(
            rf"(\d[\d,]*(?:\.\d+)?)({_LIMIT_UNIT_PATTERN})",
            table_match.group(1),
        )
        if len(table_values) == len(ordered_codes):
            for code, (raw_amount, unit) in zip(
                ordered_codes, table_values, strict=True
            ):
                limits[code] = _limit_amount(raw_amount, unit)
        else:
            # PDF tables often lose their column boundaries during text
            # extraction, for example ``10.0010.0010.00``.  Two-decimal
            # values and ``-`` placeholders still preserve the column order.
            table_tokens = re.findall(
                r"\d[\d,]*?\.\d{2}|-",
                table_match.group(1),
            )
            if len(table_tokens) >= len(ordered_codes):
                for code, raw_amount in zip(
                    ordered_codes,
                    table_tokens[: len(ordered_codes)],
                    strict=True,
                ):
                    if raw_amount != "-":
                        limits[code] = Decimal(raw_amount.replace(",", ""))

    common_limit = re.search(
        r"(?<!下属分级基金的)(?<!下属基金份额的)"
        r"限制申购(?:（[^）]*）)?金额"
        rf"[（(]单位[：:]?(?P<unit>{_LIMIT_UNIT_PATTERN})[）)]"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)",
        compact,
    )
    if common_limit:
        currency = _limit_currency(common_limit.group("unit"))
        amount = _limit_amount(
            common_limit.group("amount"), common_limit.group("unit")
        )
        currency_shares = [
            share for share in shares.values() if share.currency == currency
        ]
        # A whitespace-free PDF table can look like one common amount even
        # though it contains one value per share. Do not spread it when an
        # explicit multi-share code table is present.
        if len(currency_shares) == 1 or not ordered_codes:
            for share in currency_shares:
                limits.setdefault(share.code, amount)

    for match in re.finditer(
        r"(?P<label>(?:本基金)?(?:[A-Z]类[、，及和]?)+基金份额)"
        r".{0,40}?(?:限制金额|限额)(?:为)?"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)"
        rf"(?P<unit>{_LIMIT_UNIT_PATTERN})",
        compact,
    ):
        classes = set(re.findall(r"([A-Z])类", match.group("label")))
        currency = _limit_currency(match.group("unit"))
        amount = _limit_amount(match.group("amount"), match.group("unit"))
        for share in shares.values():
            if share.currency == currency and share.share_class in classes:
                limits[share.code] = amount

    scoped_limit = re.search(
        r"(?:该基金份额的限制金额|调整后限额)(?:为)?"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)"
        rf"(?P<unit>{_LIMIT_UNIT_PATTERN})",
        compact,
    )
    if scoped_limit:
        currency = _limit_currency(scoped_limit.group("unit"))
        amount = _limit_amount(
            scoped_limit.group("amount"), scoped_limit.group("unit")
        )
        for share in shares.values():
            if share.currency == currency:
                limits.setdefault(share.code, amount)

    cumulative_limit = re.search(
        r"(?:单日)?累计(?:申购及定期定额投资)?金额"
        r"(?:应不超过|不应超过|不超过|不高于|超过|为)?"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)"
        rf"(?P<unit>{_LIMIT_UNIT_PATTERN})(?:以下)?",
        compact,
    )
    if cumulative_limit and len(shares) == 1:
        currency = _limit_currency(cumulative_limit.group("unit"))
        amount = _limit_amount(
            cumulative_limit.group("amount"), cumulative_limit.group("unit")
        )
        for share in shares.values():
            if share.currency == currency:
                limits.setdefault(share.code, amount)

    per_share_account_limit = re.search(
        r"单个基金账户单日累计申购"
        r"(?:（含定期定额投资）)?"
        r"单个基金份额的金额"
        r"(?:不得超过|不超过|不高于)"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)"
        rf"(?P<unit>{_LIMIT_UNIT_PATTERN})",
        compact,
    )
    if per_share_account_limit:
        currency = _limit_currency(per_share_account_limit.group("unit"))
        amount = _limit_amount(
            per_share_account_limit.group("amount"),
            per_share_account_limit.group("unit"),
        )
        for share in shares.values():
            if share.currency == currency:
                limits.setdefault(share.code, amount)

    for match in re.finditer(
        r"(?P<label>"
        r"(?:申购)?本基金[^，。；]{0,80}?份额|"
        r"单个(?:人民币|美元)份额类别|"
        r"(?:人民币|美元)份额（基金代码：[^）]+(?:（[^）]+）[^）]*)*）"
        r")"
        r".{0,260}?"
        r"(?:业务限额为|累计限额为|累计金额应不超过|累计高于)"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)"
        rf"(?P<unit>{_LIMIT_UNIT_PATTERN})",
        compact,
    ):
        label = normalize_name(match.group("label"))
        share_classes = set(re.findall(r"([A-Z])类", label))
        currency = (
            "美元" if "美元" in label else _limit_currency(match.group("unit"))
        )
        amount = _limit_amount(match.group("amount"), match.group("unit"))
        candidates = [
            share
            for share in shares.values()
            if share.currency == currency
            and (not share_classes or share.share_class in share_classes)
        ]
        for share in candidates:
            previous = limits.get(share.code)
            limits[share.code] = amount if previous is None else min(previous, amount)
    return limits


def parse_subscription_announcement(
    title: str,
    text: str,
    shares: dict[str, SummaryShare],
    source_url: str,
) -> list[SubscriptionState]:
    compact = re.sub(r"\s+", "", text)
    normalized_title = re.sub(r"\s+", "", title)
    if "节假日" in normalized_title:
        return []
    cancels_limit = (
        ("取消" in normalized_title or "解除" in normalized_title)
        and "限制" in normalized_title
    )
    limits_large_subscription = "限制大额" in normalized_title
    pauses_large_subscription = bool(
        re.search(r"暂停.*?大额.*?申购", normalized_title)
    )
    pauses_subscription = bool(re.search(r"暂停.*?申购", normalized_title))
    adjusts_subscription_limit = bool(
        re.search(r"调整.*?申购.*?(?:上限|限额)", normalized_title)
    )
    resumes_large_subscription = bool(
        re.search(r"恢复(?:[（(]大额[）)]|大额).*?申购", normalized_title)
    )
    opens_subscription = bool(
        re.search(r"开放(?:日常)?(?:办理)?申购", normalized_title)
    )
    if cancels_limit or resumes_large_subscription:
        status = "open"
    elif pauses_subscription and not pauses_large_subscription:
        status = "suspended"
    elif (
        limits_large_subscription
        or pauses_large_subscription
        or "大额申购" in normalized_title
        or adjusts_subscription_limit
        or (
            ("申购" in normalized_title or "定投" in normalized_title)
            and any(
                token in normalized_title
                for token in ("金额", "限制", "上限", "限额")
            )
        )
    ):
        status = "limited"
    elif opens_subscription or (
        "恢复申购" in normalized_title and "暂停申购" not in normalized_title
    ):
        status = "open"
    else:
        return []

    codes = _announcement_share_codes(compact, normalized_title, shares)
    scoped_shares = {code: shares[code] for code in codes}
    limits = (
        _announcement_limits(text, compact, scoped_shares)
        if status == "limited"
        else {}
    )
    if status == "limited" and limits:
        codes = [code for code in codes if code in limits]
    effective_date = _announcement_date(compact)
    if "基金管理人直销电子交易平台" in compact:
        channel = "基金管理人直销电子交易平台"
    elif "基金管理人直销机构" in compact or "直销机构" in normalized_title:
        channel = "基金管理人直销机构"
    elif "直销柜台" in compact:
        channel = "直销柜台"
    else:
        channel = "全部渠道"
    return [
        SubscriptionState(
            code=code,
            status=status,
            limit_amount=limits.get(code),
            currency=shares[code].currency,
            effective_date=effective_date,
            source_url=source_url,
            channel=channel,
        )
        for code in codes
    ]


def subscription_documents(detail_html: str) -> list[DisclosureDocument]:
    return [
        document
        for document in disclosure_documents(detail_html, "")
        if any(
            token in document.title
            for token in ("申购", "定投", "金额", "限制", "暂停", "恢复", "大额")
        )
        and ("申购" in document.title or "定投" in document.title)
        and "节假日" not in document.title
    ][:12]


def discover_subscription_shares(
    announcements: list[tuple[DisclosureDocument, str]],
    product_name: str,
) -> dict[str, tuple[SummaryShare, str]]:
    discovered: dict[str, tuple[SummaryShare, str]] = {}
    for document, text in announcements:
        compact = re.sub(r"\s+", "", text)
        codes = _announcement_table_codes(compact)
        unit_match = re.search(
            r"金额单位((?:人民币元|美元)+)下属基金份额的限制申购金额",
            compact,
        )
        units = (
            re.findall(r"人民币元|美元", unit_match.group(1))
            if unit_match
            else []
        )
        name_block_match = re.search(
            r"下属(?:分级)?基金(?:份额)?的(?:基金)?简称(.+?)"
            r"下属(?:分级)?基金(?:份额)?的交易代码",
            compact,
        )
        explicit_names: list[str] = []
        name_suffixes: list[str] = []
        if name_block_match:
            name_block = name_block_match.group(1)
            name_bases = (
                product_name,
                product_name.replace("发起式", ""),
                product_name.replace("发起", ""),
            )
            for name_base in dict.fromkeys(name_bases):
                name_parts = name_block.split(name_base)
                if len(name_parts) - 1 == len(codes):
                    name_suffixes = name_parts[1:]
                    explicit_names = [
                        f"{name_base}{suffix}" for suffix in name_suffixes
                    ]
                    break
        for index, code in enumerate(codes):
            explicit_display_name = (
                explicit_names[index]
                if index < len(explicit_names)
                else None
            )
            contexts = re.findall(
                rf"([^，。；]{{0,100}}?)(?:基金)?份额[（(]基金代码[：:]"
                rf"{code}[）)]",
                compact,
            )
            context = (
                name_suffixes[index]
                if index < len(name_suffixes)
                else contexts[-1]
                if contexts
                else ""
            )
            share_class_match = re.search(r"([A-Z])类", context.upper())
            if share_class_match is None:
                share_class_match = re.search(r"([A-Z])", context.upper())
            share_class = (
                share_class_match.group(1) if share_class_match else None
            )
            _, currency, currency_form = share_identity(context)
            if index < len(units):
                currency = "美元" if units[index] == "美元" else "人民币"
            currency_annotation = re.search(
                rf"{code}[（(](人民币|美元(?:现汇|现钞)?)[）)]",
                compact,
            )
            if currency_annotation:
                annotation = currency_annotation.group(1)
                currency = "美元" if annotation.startswith("美元") else "人民币"
                currency_form = (
                    annotation.removeprefix("美元") or None
                    if currency == "美元"
                    else None
                )
            suffix = share_class or ""
            if currency == "美元":
                suffix += f"（美元{currency_form or ''}）"
            display_name = explicit_display_name or f"{product_name}{suffix}"
            if currency == "美元" and "美元" not in normalize_name(display_name):
                display_name += f"（美元{currency_form or ''}）"
            merge_summary_share(
                discovered,
                SummaryShare(
                    code,
                    display_name,
                    share_class,
                    currency,
                    currency_form,
                ),
                document.url,
            )
    return discovered


def current_announcement_shares(
    announcement_shares: dict[str, tuple[SummaryShare, str]],
    current_codes: set[str],
) -> dict[str, tuple[SummaryShare, str]]:
    """Use historical announcements only to enrich currently evidenced shares."""

    return {
        code: source
        for code, source in announcement_shares.items()
        if code in current_codes
    }


def merge_summary_share(
    shares: dict[str, tuple[SummaryShare, str]],
    share: SummaryShare,
    source_url: str,
) -> None:
    """Keep the most specific official share name found across disclosures."""

    def specificity(item: SummaryShare) -> tuple[int, int, int, int]:
        compact_name = normalize_name(item.display_name)
        return (
            int(item.share_class is not None),
            int(item.currency != "人民币" or "人民币" in compact_name),
            int(item.currency_form is not None),
            len(compact_name) if item.display_name != item.code else 0,
        )

    existing = shares.get(share.code)
    if existing is None or specificity(share) > specificity(existing[0]):
        shares[share.code] = (share, source_url)


def fetch_subscription_announcements(
    client: httpx.Client,
    detail_html: str,
) -> tuple[list[tuple[DisclosureDocument, str]], list[str]]:
    announcements: list[tuple[DisclosureDocument, str]] = []
    warnings: list[str] = []
    for document in subscription_documents(detail_html):
        try:
            announcements.append(
                (document, fetch_disclosure_text(client, document))
            )
        except (httpx.HTTPError, RuntimeError, ValueError) as exc:
            warnings.append(f"{document.title}: {exc}")
    return announcements, warnings


def resolve_subscription_states(
    announcements: list[tuple[DisclosureDocument, str]],
    shares: dict[str, SummaryShare],
    snapshot_date: date,
    detail_url: str,
) -> tuple[dict[str, SubscriptionState], list[str]]:
    states: dict[str, SubscriptionState] = {}
    warnings: list[str] = []
    for document, text in announcements:
        try:
            for state in parse_subscription_announcement(
                document.title, text, shares, document.url
            ):
                if state.effective_date > snapshot_date:
                    continue
                current = states.get(state.code)
                if current is None or state.effective_date > current.effective_date:
                    states[state.code] = state
        except (httpx.HTTPError, RuntimeError, ValueError) as exc:
            warnings.append(f"{document.title}: {exc}")

    return states, warnings


def subscription_as_of_date(collected_at: datetime) -> date:
    """Use the collection date for announcements, independent of lagging NAVs."""
    return collected_at.astimezone(ASIA_SHANGHAI).date()


def sync_subscription_state(
    session: Session,
    share_id: int,
    state: SubscriptionState,
    collected_at: datetime,
    *,
    authoritative: bool = False,
) -> None:
    effective_from = datetime.combine(
        state.effective_date, time.min, tzinfo=ASIA_SHANGHAI
    )
    records = list(
        session.scalars(
            select(SalesLimitHistory)
            .where(
                SalesLimitHistory.fund_share_class_id == share_id,
                SalesLimitHistory.investor_type == "全部投资者",
                SalesLimitHistory.business_type == "申购",
            )
            .order_by(
                SalesLimitHistory.effective_from.desc(),
                SalesLimitHistory.id.desc(),
            )
        )
    )
    effective_limit_amount = state.limit_amount
    if effective_limit_amount is None:
        # A later parser run must not discard a verified amount from the
        # exact same official document and effective state. Reuse the richer
        # audited value and allow authoritative sync to reopen that record.
        richer_record = next(
            (
                record
                for record in records
                if record.channel == state.channel
                and record.effective_from == effective_from
                and record.limit_status == state.status
                and record.currency == state.currency
                and record.source_url == state.source_url
                and record.limit_amount is not None
            ),
            None,
        )
        if richer_record is not None:
            effective_limit_amount = richer_record.limit_amount
    active_records = [record for record in records if record.effective_to is None]
    matching_records = [
        record
        for record in records
        if record.channel == state.channel
        and record.effective_from == effective_from
        and record.limit_status == state.status
        and record.limit_amount == effective_limit_amount
        and record.currency == state.currency
    ]
    existing = next(
        (record for record in matching_records if record.effective_to is None),
        None,
    )
    if existing is None and authoritative and matching_records:
        existing = matching_records[0]
        existing.effective_to = None

    newer_active_exists = any(
        record.effective_from is not None
        and record.effective_from > effective_from
        for record in active_records
    )
    if newer_active_exists and not authoritative:
        return

    for record in active_records:
        if record is existing:
            continue
        if record.effective_from is None or record.effective_from <= effective_from:
            record.effective_to = max(
                effective_from,
                record.effective_from or effective_from,
            )
        elif authoritative:
            # The complete official announcement timeline no longer assigns
            # this later record to the share (for example, a RMB-only notice
            # previously applied to a USD share). Keep the row for audit, but
            # make its effective interval empty so it cannot remain current.
            record.effective_to = record.effective_from
    if existing is not None:
        existing.source_url = state.source_url
        existing.source_time = effective_from
        existing.collected_at = collected_at
        existing.quality_status = state.quality_status
        return
    session.add(
        SalesLimitHistory(
            fund_share_class_id=share_id,
            channel=state.channel,
            investor_type="全部投资者",
            business_type="申购",
            limit_amount=effective_limit_amount,
            currency=state.currency,
            limit_status=state.status,
            source_url=state.source_url,
            source_time=effective_from,
            effective_from=effective_from,
            collected_at=collected_at,
            quality_status=state.quality_status,
        )
    )


def _quarter_end(year: int, quarter: int) -> date:
    return date(year, quarter * 3, 31 if quarter in (1, 4) else 30)


def _report_year(value: str) -> int:
    digits = value.translate(
        str.maketrans("〇零一二三四五六七八九", "00123456789")
    )
    return int(digits)


def parse_quarterly_scales(text: str) -> list[ScaleRecord]:
    compact = re.sub(r"\s+", "", text)
    period = re.search(
        r"([0-9〇零一二三四五六七八九]{4})年第?([一二三四1-4])季度报告",
        compact,
    )
    if period is None:
        raise RuntimeError("Quarterly report did not contain a report period")
    quarter_map = {"一": 1, "二": 2, "三": 3, "四": 4}
    raw_quarter = period.group(2)
    quarter = quarter_map.get(raw_quarter, int(raw_quarter) if raw_quarter.isdigit() else 0)
    report_date = _quarter_end(_report_year(period.group(1)), quarter)

    codes_match = re.search(
        r"下属(?:分级)?基金的交易代码((?:\d{6})+)", compact
    )
    if codes_match:
        share_codes = _code_chunks(codes_match.group(1))
    else:
        code_match = re.search(
            r"(?:基金主代码|基金代码|交易代码)(\d{6})",
            compact,
        )
        if code_match is None:
            raise RuntimeError("Quarterly report did not contain share codes")
        share_codes = (code_match.group(1),)

    values_match = re.search(
        r"(?:4\.)?期末基金资产净值(.+?)"
        r"(?:5\.期末基金份.*?额净值|期末基金份额净值)", compact
    )
    if values_match is None:
        raise RuntimeError("Quarterly report did not contain ending net assets")
    values = re.findall(r"(?:\d[\d,]*\.\d{2}|-)", values_match.group(1))
    if len(values) < len(share_codes):
        raise RuntimeError("Quarterly report share codes and net assets did not align")
    return [
        ScaleRecord(code, report_date, Decimal(raw.replace(",", "")))
        for code, raw in zip(share_codes, values, strict=False)
        if raw != "-"
    ]


def fetch_max_valuation_date(client: httpx.Client) -> date:
    response = client.post(EID_SEARCH_META_URL, data={"reportTypeStatus": "01"})
    response.raise_for_status()
    value = _as_date(response.json().get("maxValuationDate"))
    if value is None:
        raise RuntimeError("EID did not provide maxValuationDate")
    return value


def discovery_name(name: str) -> str:
    value = name.strip()
    value = re.sub(
        r"([A-Z])(?=(?:人民币|美元(?:现汇|现钞)?)?(?:\s*\([^)]*\))?$)",
        "",
        value,
        flags=re.I,
    )
    value = re.sub(r"(?:人民币|美元(?:现汇|现钞)?)\s*$", "", value)
    return value.strip()


def nav_query_data(
    name: str,
    start_date: date,
    end_date: date,
    *,
    fund_code: str = "",
) -> str:
    data = [
        {"name": "sEcho", "value": 1},
        {"name": "iColumns", "value": 5},
        {"name": "sColumns", "value": ""},
        {"name": "iDisplayStart", "value": 0},
        {"name": "iDisplayLength", "value": 500},
        {"name": "fundType", "value": "all"},
        {"name": "fundCompanyShortName", "value": ""},
        {"name": "fundCode", "value": fund_code},
        {"name": "fundName", "value": quote(discovery_name(name), safe="")},
        {"name": "startDate", "value": start_date.isoformat()},
        {"name": "endDate", "value": end_date.isoformat()},
    ]
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def _decimal(value: Any) -> Decimal | None:
    try:
        result = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    return result if result > 0 else None


def share_identity(display_name: str) -> tuple[str | None, str, str | None]:
    compact = normalize_name(display_name).replace("(", "").replace(")", "")
    currency = "美元" if any(token in compact for token in ("美元", "美汇", "美钞")) else "人民币"
    currency_form = (
        "现汇"
        if any(token in compact for token in ("现汇", "美汇"))
        else "现钞"
        if any(token in compact for token in ("现钞", "美钞"))
        else None
    )
    for pattern in (
        r"(?:人民币|美元(?:现汇|现钞)?)?([A-Z])$",
        r"([A-Z])(?:人民币|美元(?:现汇|现钞)?)$",
    ):
        match = re.search(pattern, compact)
        if match:
            return match.group(1), currency, currency_form
    return None, currency, currency_form


def parse_nav_rows(rows: list[dict[str, Any]], fund_id: int) -> list[NavRecord]:
    records: dict[tuple[str, date], NavRecord] = {}
    for row in rows:
        if str((row.get("fund") or {}).get("idStr")) != str(fund_id):
            continue
        code = str(row.get("code") or "").strip()
        nav_date = _as_date(row.get("valuationDate"))
        unit_nav = _decimal(row.get("shareNetValue"))
        if not re.fullmatch(r"\d{6}", code) or nav_date is None or unit_nav is None:
            continue
        display_name = str(row.get("shortName") or code).strip()
        share_class, currency, currency_form = share_identity(display_name)
        records[(code, nav_date)] = NavRecord(
            code,
            display_name,
            share_class,
            currency,
            currency_form,
            nav_date,
            unit_nav,
            _decimal(row.get("totalNetValue")),
        )
    return sorted(records.values(), key=lambda item: (item.code, item.nav_date))


def fetch_nav_records(
    client: httpx.Client,
    fund_id: int,
    name: str,
    end_date: date,
    *,
    lookback_days: int = 45,
    fund_code: str = "",
) -> list[NavRecord]:
    rows: list[dict[str, Any]] = []
    window_start = end_date - timedelta(days=lookback_days)
    while window_start <= end_date:
        # The endpoint returns an empty result for a roughly one-year query,
        # even when matching rows exist. Keep each request within the proven
        # working recent-NAV range and merge the windows locally.
        window_end = min(window_start + timedelta(days=45), end_date)
        payload: dict[str, Any] = {}
        for attempt in range(3):
            response = client.get(
                EID_NAV_URL,
                params={
                    "aoData": nav_query_data(
                        name,
                        window_start,
                        window_end,
                        fund_code=fund_code,
                    )
                },
            )
            response.raise_for_status()
            payload = response.json()
            message = str(payload.get("message") or "")
            if payload.get("success") is not False:
                break
            if "请求繁忙" not in message or attempt == 2:
                raise RuntimeError(message or "EID NAV query failed")
        rows.extend(payload.get("aaData") or [])
        window_start = window_end + timedelta(days=1)
    return parse_nav_rows(rows, fund_id)


def nav_lookback_days(
    session: Session,
    product_id: int,
    end_date: date,
) -> int:
    earliest = session.scalar(
        select(NavDaily.nav_date)
        .join(
            FundShareClass,
            FundShareClass.id == NavDaily.fund_share_class_id,
        )
        .where(FundShareClass.fund_product_id == product_id)
        .order_by(NavDaily.nav_date)
        .limit(1)
    )
    return 400 if earliest is None or earliest > end_date - timedelta(days=370) else 45


def share_nav_lookback_days(
    session: Session,
    code: str,
    end_date: date,
) -> int:
    earliest = session.scalar(
        select(NavDaily.nav_date)
        .join(
            FundShareClass,
            FundShareClass.id == NavDaily.fund_share_class_id,
        )
        .where(FundShareClass.code == code)
        .order_by(NavDaily.nav_date)
        .limit(1)
    )
    return 400 if earliest is None or earliest > end_date - timedelta(days=370) else 45


def recent_nav_share_codes(
    session: Session,
    product_id: int,
    end_date: date,
    *,
    max_age_days: int = 45,
) -> set[str]:
    return set(
        session.scalars(
            select(FundShareClass.code)
            .join(NavDaily, NavDaily.fund_share_class_id == FundShareClass.id)
            .where(
                FundShareClass.fund_product_id == product_id,
                NavDaily.nav_date >= end_date - timedelta(days=max_age_days),
            )
            .distinct()
        )
    )


def current_product_share_codes(
    candidate_code: str,
    summary_shares: dict[str, tuple[SummaryShare, str]],
    recent_nav_codes: set[str],
) -> set[str]:
    return {candidate_code, *summary_shares, *recent_nav_codes}


def retire_stale_product_shares(
    session: Session,
    product_id: int,
    current_codes: set[str],
    collected_at: datetime,
) -> int:
    stale_share_ids = list(
        session.scalars(
            select(FundShareClass.id).where(
                FundShareClass.fund_product_id == product_id,
                FundShareClass.status == "active",
                FundShareClass.code.not_in(sorted(current_codes)),
            )
        )
    )
    if not stale_share_ids:
        return 0
    session.execute(
        update(SalesLimitHistory)
        .where(
            SalesLimitHistory.fund_share_class_id.in_(stale_share_ids),
            SalesLimitHistory.effective_to.is_(None),
        )
        .values(effective_to=collected_at)
    )
    session.execute(
        update(FundShareClass)
        .where(FundShareClass.id.in_(stale_share_ids))
        .values(status="inactive", updated_at=collected_at)
    )
    return len(stale_share_ids)


def sync_fund(
    session: Session,
    candidate: ProductCandidate,
    detail: FundDetail,
    nav_records: list[NavRecord],
    summary_shares: dict[str, tuple[SummaryShare, str]],
    snapshot_time: datetime,
    collected_at: datetime,
    benchmark_description: str | None = None,
) -> tuple[int, dict[str, int], int]:
    canonical_code = f"csrc:{detail.fund_id}"
    product_values = {
        "registration_code": candidate.code,
        "name": detail.full_name,
        "fund_company": detail.manager,
        "product_structure": candidate.product_structure,
        "trading_venue": "仅场外",
        "investment_scopes": list(candidate.target.investment_scopes),
        "tracking_method": "被动指数",
        "exact_benchmark_id": candidate.target.definition_id,
        "benchmark_description": (
            benchmark_description or candidate.target.benchmark_description
        ),
        "inception_date": detail.inception_date or candidate.inception_date,
        "status": "active",
        "source_url": f"{EID_DETAIL_URL}?fundId={detail.fund_id}",
        "source_time": snapshot_time,
        "effective_from": snapshot_time,
        "collected_at": collected_at,
        "quality_status": "verified",
    }
    statement = insert(FundProduct).values(
        canonical_code=canonical_code, **product_values
    )
    session.execute(
        statement.on_conflict_do_update(
            constraint="uq_fund_product_canonical_code",
            set_={**product_values, "updated_at": collected_at},
        )
    )
    product_id = session.scalar(
        select(FundProduct.id).where(FundProduct.canonical_code == canonical_code)
    )
    if product_id is None:
        raise RuntimeError(f"Product upsert failed for {candidate.code}")

    share_ids: dict[str, int] = {}
    latest_by_code: dict[str, NavRecord] = {}
    for record in nav_records:
        previous = latest_by_code.get(record.code)
        if previous is None or record.nav_date > previous.nav_date:
            latest_by_code[record.code] = record

    share_codes = set(latest_by_code) | set(summary_shares)
    for code in sorted(share_codes):
        record = latest_by_code.get(code)
        summary_source = summary_shares.get(code)
        # NAV rows occasionally expose only the product-level short name (for
        # example, 050025 without its A suffix). Product summaries are the
        # authoritative source for share-class identity, so prefer their
        # explicit name/metadata whenever available and still sync the NAV row
        # separately below.
        if (
            summary_source is not None
            and summary_source[0].display_name != code
        ):
            summary_share, source_url = summary_source
            display_name = summary_share.display_name
            share_class = summary_share.share_class
            currency = summary_share.currency
            currency_form = summary_share.currency_form
            source_time = snapshot_time
            quality_status = "verified"
        elif record is not None:
            display_name = record.display_name
            share_class = record.share_class
            currency = record.currency
            currency_form = record.currency_form
            source_url = record.source_url
            source_time = datetime.combine(
                record.nav_date, time.min, tzinfo=ASIA_SHANGHAI
            )
            quality_status = "verified"
        elif summary_source is not None:
            summary_share, source_url = summary_source
            display_name = summary_share.display_name
            if display_name == code:
                currency_label = (
                    f"（{summary_share.currency}{summary_share.currency_form or ''}）"
                    if summary_share.currency != "人民币"
                    else ""
                )
                display_name = (
                    f"{detail.short_name}{summary_share.share_class or ''}"
                    f"{currency_label}"
                )
            share_class = summary_share.share_class
            currency = summary_share.currency
            currency_form = summary_share.currency_form
            source_time = snapshot_time
            quality_status = "verified"
        else:
            continue
        share_values = {
            "fund_product_id": product_id,
            "display_name": display_name,
            "share_class": share_class,
            "currency": currency,
            "currency_form": currency_form,
            "inception_date": candidate.inception_date if code == candidate.code else None,
            "status": "active",
            "source_url": source_url,
            "source_time": source_time,
            "effective_from": snapshot_time,
            "collected_at": collected_at,
            "quality_status": quality_status,
        }
        share_statement = insert(FundShareClass).values(code=code, **share_values)
        session.execute(
            share_statement.on_conflict_do_update(
                index_elements=[FundShareClass.code],
                set_={**share_values, "updated_at": collected_at},
            )
        )
        share_id = session.scalar(
            select(FundShareClass.id).where(FundShareClass.code == code)
        )
        if share_id is None:
            raise RuntimeError(f"Share upsert failed for {code}")
        share_ids[code] = share_id

    for record in nav_records:
        values = {
            "unit_nav": record.unit_nav,
            "accumulated_nav": record.accumulated_nav,
            "source_url": record.source_url,
            "source_time": datetime.combine(
                record.nav_date, time.min, tzinfo=ASIA_SHANGHAI
            ),
            "effective_from": datetime.combine(
                record.nav_date, time.min, tzinfo=ASIA_SHANGHAI
            ),
            "collected_at": collected_at,
            "quality_status": "verified",
        }
        nav_statement = insert(NavDaily).values(
            fund_share_class_id=share_ids[record.code],
            nav_date=record.nav_date,
            **values,
        )
        session.execute(
            nav_statement.on_conflict_do_update(
                constraint="uq_nav_daily_share_date", set_=values
            )
        )
    return product_id, share_ids, len(nav_records)


def sync_scale_history(
    session: Session,
    share_id: int,
    record: ScaleRecord,
    collected_at: datetime,
    source_url: str,
) -> None:
    source_time = datetime.combine(
        record.report_date, time.min, tzinfo=ASIA_SHANGHAI
    )
    existing = session.scalar(
        select(FundScale)
        .where(
            FundScale.fund_share_class_id == share_id,
            FundScale.report_date == record.report_date,
        )
        .order_by(FundScale.id.desc())
        .limit(1)
    )
    values = {
        "amount": record.amount_cny,
        "amount_cny": record.amount_cny,
        "currency": "人民币",
        "source_url": source_url,
        "source_time": source_time,
        "effective_from": source_time,
        "collected_at": collected_at,
        "quality_status": "verified",
    }
    if existing is None:
        session.add(
            FundScale(
                fund_share_class_id=share_id,
                report_date=record.report_date,
                **values,
            )
        )
        return
    for key, value in values.items():
        setattr(existing, key, value)


def sync_catalog_candidate(
    session: Session,
    candidate: ProductCandidate,
    detail: FundDetail,
    snapshot_time: datetime,
    collected_at: datetime,
    source_url: str,
) -> tuple[int, int]:
    product_values = {
        "registration_code": candidate.code,
        "name": detail.full_name,
        "fund_company": detail.manager,
        "product_structure": candidate.product_structure,
        "trading_venue": "仅场外",
        "investment_scopes": list(candidate.target.investment_scopes),
        "tracking_method": "被动指数",
        "exact_benchmark_id": candidate.target.definition_id,
        "benchmark_description": candidate.target.benchmark_description,
        "inception_date": detail.inception_date or candidate.inception_date,
        "status": "active",
        "source_url": source_url,
        "source_time": snapshot_time,
        "effective_from": snapshot_time,
        "collected_at": collected_at,
        "quality_status": "verified",
    }
    canonical_code = f"csrc:{detail.fund_id}"
    statement = insert(FundProduct).values(
        canonical_code=canonical_code,
        **product_values,
    )
    session.execute(
        statement.on_conflict_do_update(
            constraint="uq_fund_product_canonical_code",
            set_={**product_values, "updated_at": collected_at},
        )
    )
    product_id = session.scalar(
        select(FundProduct.id).where(
            FundProduct.canonical_code == canonical_code
        )
    )
    if product_id is None:
        raise RuntimeError(f"Catalog product upsert failed for {candidate.code}")

    display_name = detail.short_name or candidate.name
    share_class, currency, currency_form = share_identity(display_name)
    share_values = {
        "fund_product_id": product_id,
        "display_name": display_name,
        "share_class": share_class,
        "currency": currency,
        "currency_form": currency_form,
        "inception_date": detail.inception_date or candidate.inception_date,
        "status": "active",
        "source_url": source_url,
        "source_time": snapshot_time,
        "effective_from": snapshot_time,
        "collected_at": collected_at,
        "quality_status": "verified",
    }
    share_statement = insert(FundShareClass).values(
        code=candidate.code,
        **share_values,
    )
    # Script E owns the low-frequency catalog/master-data lifecycle. When a
    # share already exists, do not replace the richer A/C/E and currency
    # identity populated by script F with the product-level catalog short name.
    catalog_share_updates = {
        key: value
        for key, value in share_values.items()
        if key
        not in {
            "display_name",
            "share_class",
            "currency",
            "currency_form",
        }
    }
    session.execute(
        share_statement.on_conflict_do_update(
            index_elements=[FundShareClass.code],
            set_={**catalog_share_updates, "updated_at": collected_at},
        )
    )
    share_id = session.scalar(
        select(FundShareClass.id).where(FundShareClass.code == candidate.code)
    )
    if share_id is None:
        raise RuntimeError(f"Catalog share upsert failed for {candidate.code}")
    return product_id, share_id


def should_retire_stale_catalog_products(
    *,
    failures: list[str],
    codes: tuple[str, ...] | None,
    limit: int | None,
) -> bool:
    """Only a successful, complete catalogue run may retire missing products."""

    return not failures and not codes and limit is None


def run_sync(
    *,
    dry_run: bool = False,
    limit: int | None = None,
    codes: tuple[str, ...] | None = None,
) -> CatalogSyncStats:
    headers = {"User-Agent": "index-fund-comparator/0.1"}
    transport = httpx.HTTPTransport(retries=2)
    collected_at = datetime.now(UTC)
    with httpx.Client(
        timeout=httpx.Timeout(40, connect=10),
        headers=headers,
        transport=transport,
    ) as client:
        content, index_url = fetch_product_index(client)
        candidates, snapshot_date = parse_product_index(
            content,
            fallback_snapshot_date=snapshot_date_from_url(index_url),
        )
        if codes:
            requested_codes = {
                code.strip().zfill(6) for code in codes if code.strip()
            }
            candidates = [
                candidate
                for candidate in candidates
                if candidate.code in requested_codes
            ]
            found_codes = {candidate.code for candidate in candidates}
            missing_codes = sorted(requested_codes - found_codes)
            if missing_codes:
                raise RuntimeError(
                    "Requested codes were not eligible index candidates: "
                    + ", ".join(missing_codes)
                )
        if limit is not None:
            candidates = candidates[:limit]

        snapshot_time = datetime.combine(
            snapshot_date,
            time.min,
            tzinfo=ASIA_SHANGHAI,
        )
        products = shares = fee_shares = scales = 0
        active_product_ids: list[int] = []
        failures: list[str] = []
        warnings: list[str] = []
        with get_session_factory()() as session:
            ensure_index_master_data(session, snapshot_time, source_url=index_url)
            for candidate in candidates:
                try:
                    with session.begin_nested():
                        fund_id = validate_fund_code(client, candidate.code)
                        detail, detail_html = fetch_fund_detail_document(
                            client, fund_id
                        )
                        classified = classify_product(detail.full_name)
                        if classified is None:
                            continue
                        target, structure = classified
                        resolved_candidate = ProductCandidate(
                            candidate.code,
                            candidate.name,
                            candidate.inception_date,
                            target,
                            structure,
                        )

                        exchange_traded: bool | None = None
                        summary_discovery_complete = True
                        summaries: list[
                            tuple[DisclosureDocument, ProductSummary]
                        ] = []
                        try:
                            summary_documents = product_summary_documents(
                                client,
                                detail_html,
                                candidate.code,
                                collected_at.date(),
                            )
                        except (httpx.HTTPError, RuntimeError, ValueError) as exc:
                            summary_discovery_complete = False
                            warnings.append(
                                f"{candidate.code}: advanced summary search: {exc}"
                            )
                            summary_documents = disclosure_documents(
                                detail_html, PRODUCT_SUMMARY_LABEL
                            )
                        for document in summary_documents:
                            try:
                                summary_text = fetch_disclosure_text(
                                    client, document
                                )
                                if exchange_traded is None:
                                    exchange_traded = is_exchange_traded_summary(
                                        summary_text
                                    )
                                summaries.append(
                                    (
                                        document,
                                        parse_product_summary(summary_text),
                                    )
                                )
                            except (
                                httpx.HTTPError,
                                RuntimeError,
                                ValueError,
                            ) as exc:
                                summary_discovery_complete = False
                                warnings.append(
                                    f"{candidate.code} {document.title}: {exc}"
                                )
                        if exchange_traded:
                            continue
                        if not summaries:
                            summary_discovery_complete = False
                            warnings.append(
                                f"{candidate.code} {candidate.name}: "
                                "no usable product summary"
                            )

                        scale_document: DisclosureDocument | None = None
                        scale_records: list[ScaleRecord] = []
                        quarterly_documents = disclosure_documents(
                            detail_html, QUARTERLY_REPORT_LABEL
                        )
                        if quarterly_documents:
                            scale_document = quarterly_documents[0]
                            try:
                                report_text = fetch_disclosure_text(
                                    client, scale_document
                                )
                                scale_records = parse_quarterly_scales(
                                    report_text
                                )
                            except (
                                httpx.HTTPError,
                                RuntimeError,
                                ValueError,
                            ) as exc:
                                warnings.append(
                                    f"{candidate.code} "
                                    f"{scale_document.title}: {exc}"
                                )
                        else:
                            warnings.append(
                                f"{candidate.code} {candidate.name}: "
                                "no quarterly report"
                            )

                        product_id, primary_share_id = sync_catalog_candidate(
                            session,
                            resolved_candidate,
                            detail,
                            snapshot_time,
                            collected_at,
                            fund_detail_url(fund_id),
                        )
                        summary_shares: dict[
                            str, tuple[SummaryShare, str]
                        ] = {}
                        for document, summary in summaries:
                            for summary_share in summary.shares:
                                merge_summary_share(
                                    summary_shares,
                                    summary_share,
                                    document.url,
                                )
                        benchmark = next(
                            (
                                summary.benchmark_description
                                for _, summary in summaries
                                if summary.benchmark_description
                            ),
                            None,
                        )
                        _, discovered_share_ids, _ = sync_fund(
                            session,
                            resolved_candidate,
                            detail,
                            [],
                            summary_shares,
                            snapshot_time,
                            collected_at,
                            benchmark,
                        )
                        share_ids = {
                            candidate.code: primary_share_id,
                            **discovered_share_ids,
                        }
                        if summary_discovery_complete:
                            current_codes = current_product_share_codes(
                                candidate.code,
                                summary_shares,
                                recent_nav_share_codes(
                                    session, product_id, snapshot_date
                                ),
                            )
                            retire_stale_product_shares(
                                session,
                                product_id,
                                current_codes,
                                collected_at,
                            )

                        rates_by_code: dict[
                            str, tuple[dict[str, Decimal], str]
                        ] = {}
                        for document, summary in summaries:
                            for code in summary.share_codes:
                                rates_by_code.setdefault(
                                    code,
                                    (
                                        summary.rates_by_code.get(
                                            code, summary.rates
                                        ),
                                        document.url,
                                    ),
                                )
                        common_rates: (
                            tuple[dict[str, Decimal], str] | None
                        ) = None
                        if summaries:
                            document, summary = summaries[0]
                            common_rates = (
                                {
                                    fee_type: rate
                                    for fee_type, rate in summary.rates.items()
                                    if fee_type in {"management", "custody"}
                                },
                                document.url,
                            )
                        synced_fee_shares = 0
                        for code, share_id in share_ids.items():
                            rate_source = rates_by_code.get(code) or common_rates
                            if rate_source is None:
                                continue
                            rates, source_url = rate_source
                            sync_fee_history(
                                session,
                                share_id,
                                rates,
                                collected_at,
                                source_url,
                            )
                            synced_fee_shares += 1

                        synced_scales = 0
                        for scale_record in scale_records:
                            share_id = share_ids.get(scale_record.share_code)
                            if share_id is None:
                                warnings.append(
                                    f"{candidate.code}: quarterly scale share "
                                    f"{scale_record.share_code} was not discovered"
                                )
                                continue
                            assert scale_document is not None
                            sync_scale_history(
                                session,
                                share_id,
                                scale_record,
                                collected_at,
                                scale_document.url,
                            )
                            synced_scales += 1
                        active_product_ids.append(product_id)
                    products += 1
                    shares += len(share_ids)
                    fee_shares += synced_fee_shares
                    scales += synced_scales
                except (httpx.HTTPError, RuntimeError, ValueError) as exc:
                    failures.append(f"{candidate.code} {candidate.name}: {exc}")
            if products == 0:
                session.rollback()
                raise RuntimeError("CSRC catalog sync produced no products")
            if should_retire_stale_catalog_products(
                failures=failures,
                codes=codes,
                limit=limit,
            ):
                stale_product_ids = list(
                    session.scalars(
                        select(FundProduct.id).where(
                            FundProduct.canonical_code.like("csrc:%"),
                            FundProduct.trading_venue == "仅场外",
                            FundProduct.status == "active",
                            FundProduct.id.not_in(active_product_ids),
                        )
                    )
                )
                if stale_product_ids:
                    session.execute(
                        update(FundProduct)
                        .where(FundProduct.id.in_(stale_product_ids))
                        .values(status="inactive", updated_at=collected_at)
                    )
                    session.execute(
                        update(FundShareClass)
                        .where(
                            FundShareClass.fund_product_id.in_(stale_product_ids)
                        )
                        .values(status="inactive", updated_at=collected_at)
                    )
            if dry_run:
                session.rollback()
            else:
                session.commit()
    return CatalogSyncStats(
        products,
        shares,
        snapshot_date,
        tuple(failures),
        fee_shares,
        scales,
        tuple(warnings),
    )


def run_details_sync(
    *,
    dry_run: bool = False,
    limit: int | None = None,
    codes: tuple[str, ...] | None = None,
) -> SyncStats:
    headers = {"User-Agent": "index-fund-comparator/0.1"}
    transport = httpx.HTTPTransport(retries=2)
    with httpx.Client(
        timeout=httpx.Timeout(40, connect=10),
        headers=headers,
        transport=transport,
    ) as client:
        company_nav_fetcher = OfficialCompanyNavFetcher(client)
        max_nav_date = fetch_max_valuation_date(client)
        snapshot_date = max_nav_date
        snapshot_time = datetime.combine(
            snapshot_date, time.min, tzinfo=ASIA_SHANGHAI
        )
        collected_at = datetime.now(UTC)
        subscription_date = subscription_as_of_date(collected_at)
        products = shares = nav_rows = fee_shares = scales = subscription_states = 0
        return_metrics = 0
        failures: list[str] = []
        warnings: list[str] = []
        with get_session_factory()() as session:
            ensure_index_master_data(
                session,
                snapshot_time,
                source_url=EID_SEARCH_META_URL,
            )
            stored_products = list(
                session.scalars(
                    select(FundProduct)
                    .where(
                        FundProduct.canonical_code.like("csrc:%"),
                        FundProduct.trading_venue == "仅场外",
                        FundProduct.status == "active",
                    )
                    .order_by(FundProduct.registration_code, FundProduct.id)
                )
            )
            share_codes_by_product: dict[int, list[str]] = {}
            stored_shares_by_product: dict[int, dict[str, SummaryShare]] = {}
            for (
                product_id,
                share_code,
                display_name,
                share_class,
                currency,
                currency_form,
            ) in session.execute(
                select(
                    FundShareClass.fund_product_id,
                    FundShareClass.code,
                    FundShareClass.display_name,
                    FundShareClass.share_class,
                    FundShareClass.currency,
                    FundShareClass.currency_form,
                ).where(FundShareClass.status == "active")
            ):
                share_codes_by_product.setdefault(product_id, []).append(
                    share_code
                )
                stored_shares_by_product.setdefault(product_id, {})[
                    share_code
                ] = SummaryShare(
                    share_code,
                    display_name,
                    share_class,
                    currency,
                    currency_form,
                )

            requested_codes = {
                code.strip().zfill(6) for code in (codes or ()) if code.strip()
            }
            targets: list[tuple[FundProduct, ProductCandidate, int]] = []
            matched_codes: set[str] = set()
            for product in stored_products:
                target = TARGET_BY_DEFINITION.get(product.exact_benchmark_id or "")
                product_codes = share_codes_by_product.get(product.id, [])
                registration_code = product.registration_code or (
                    product_codes[0] if product_codes else ""
                )
                if target is None or not re.fullmatch(r"\d{6}", registration_code):
                    continue
                if requested_codes:
                    matches = requested_codes.intersection(
                        {registration_code, *product_codes}
                    )
                    if not matches:
                        continue
                    matched_codes.update(matches)
                try:
                    fund_id = int(product.canonical_code.removeprefix("csrc:"))
                except ValueError:
                    failures.append(
                        f"{registration_code} {product.name}: invalid canonical code"
                    )
                    continue
                targets.append(
                    (
                        product,
                        ProductCandidate(
                            registration_code,
                            product.name,
                            product.inception_date,
                            target,
                            product.product_structure,
                        ),
                        fund_id,
                    )
                )
            missing_codes = sorted(requested_codes - matched_codes)
            if missing_codes:
                raise RuntimeError(
                    "Requested codes were not active script E targets: "
                    + ", ".join(missing_codes)
                )
            if limit is not None:
                targets = targets[:limit]
            if not targets:
                raise RuntimeError(
                    "No active CSRC off-exchange targets found; run script E first"
                )

            for stored_product, candidate, fund_id in targets:
                try:
                    with session.begin_nested():
                        detail, detail_html = fetch_fund_detail_document(
                            client, fund_id
                        )
                        detail_classified = classify_product(detail.full_name)
                        if detail_classified is None:
                            continue
                        detail_target, detail_structure = detail_classified
                        candidate = ProductCandidate(
                            candidate.code,
                            candidate.name,
                            candidate.inception_date,
                            detail_target,
                            detail_structure,
                        )
                        lookback_days = nav_lookback_days(
                            session,
                            stored_product.id,
                            max_nav_date,
                        )
                        query_name = detail.short_name or candidate.name
                        records = fetch_nav_records(
                            client,
                            fund_id,
                            query_name,
                            max_nav_date,
                            lookback_days=lookback_days,
                        )
                        if (
                            not records
                            and discovery_name(candidate.name)
                            != discovery_name(query_name)
                        ):
                            records = fetch_nav_records(
                                client,
                                fund_id,
                                candidate.name,
                                max_nav_date,
                                lookback_days=lookback_days,
                            )
                        if not records:
                            by_identity: dict[
                                tuple[str, date], NavRecord
                            ] = {}
                            fallback_codes = (
                                share_codes_by_product.get(stored_product.id)
                                or [candidate.code]
                            )
                            for share_code in fallback_codes:
                                for record in fetch_nav_records(
                                    client,
                                    fund_id,
                                    "",
                                    max_nav_date,
                                    lookback_days=lookback_days,
                                    fund_code=share_code,
                                ):
                                    by_identity[(record.code, record.nav_date)] = record
                            records = sorted(
                                by_identity.values(),
                                key=lambda item: (item.code, item.nav_date),
                            )
                        announcements, announcement_warnings = (
                            fetch_subscription_announcements(
                                client, detail_html
                            )
                        )
                        warnings.extend(
                            f"{candidate.code} {warning}"
                            for warning in announcement_warnings
                        )
                        announcement_shares = discover_subscription_shares(
                            announcements,
                            detail.short_name or candidate.name,
                        )
                        fallback_shares = {
                            **stored_shares_by_product.get(stored_product.id, {}),
                            **{
                                code: source[0]
                                for code, source in announcement_shares.items()
                            },
                        }
                        official_records: list[NavRecord] = []
                        for share in fallback_shares.values():
                            if share.currency != "美元":
                                continue
                            official_lookback = share_nav_lookback_days(
                                session,
                                share.code,
                                max_nav_date,
                            )
                            try:
                                fallback_records = company_nav_fetcher.fetch(
                                    detail.manager,
                                    share.code,
                                    max_nav_date - timedelta(
                                        days=official_lookback
                                    ),
                                    max_nav_date,
                                )
                            except (httpx.HTTPError, RuntimeError, ValueError) as exc:
                                warnings.append(
                                    f"{share.code} {share.display_name}: "
                                    f"official company NAV fallback failed: {exc}"
                                )
                                continue
                            if not fallback_records:
                                warnings.append(
                                    f"{share.code} {share.display_name}: "
                                    "official company NAV fallback returned no rows"
                                )
                                continue
                            official_records.extend(
                                NavRecord(
                                    share.code,
                                    share.display_name,
                                    share.share_class,
                                    share.currency,
                                    share.currency_form,
                                    record.nav_date,
                                    record.unit_nav,
                                    record.accumulated_nav,
                                    record.source_url,
                                )
                                for record in fallback_records
                            )
                        if official_records:
                            by_identity = {
                                (record.code, record.nav_date): record
                                for record in official_records
                            }
                            # EID remains the primary source whenever both
                            # official sources publish the same share/date.
                            by_identity.update(
                                {
                                    (record.code, record.nav_date): record
                                    for record in records
                                }
                            )
                            records = sorted(
                                by_identity.values(),
                                key=lambda item: (item.code, item.nav_date),
                            )
                        if not records:
                            raise RuntimeError(
                                "no valid NAV rows from EID or official company sources"
                            )
                        current_codes = {
                            candidate.code,
                            *(record.code for record in records),
                            *recent_nav_share_codes(
                                session, stored_product.id, max_nav_date
                            ),
                        }
                        summary_shares: dict[
                            str, tuple[SummaryShare, str]
                        ] = {}
                        for discovered_share in current_announcement_shares(
                            announcement_shares, current_codes
                        ).values():
                            merge_summary_share(
                                summary_shares,
                                discovered_share[0],
                                discovered_share[1],
                            )
                        _, share_ids, synced_nav = sync_fund(
                            session,
                            candidate,
                            detail,
                            records,
                            summary_shares,
                            snapshot_time,
                            collected_at,
                            stored_product.benchmark_description,
                        )
                        synced_return_metrics = 0
                        metric_sources = {
                            record.code: record.source_url
                            for record in records
                        }
                        for code, share_id in share_ids.items():
                            metrics = calculate_return_metrics(
                                load_return_nav_records(session, share_id)
                            )
                            sync_return_metrics(
                                session,
                                share_id,
                                candidate.target.definition_id,
                                metrics,
                                collected_at,
                                metric_sources.get(code, EID_NAV_URL),
                            )
                            synced_return_metrics += len(metrics)

                        stored_shares = {
                            share.code: SummaryShare(
                                share.code,
                                share.display_name,
                                share.share_class,
                                share.currency,
                                share.currency_form,
                            )
                            for share in session.scalars(
                                select(FundShareClass).where(
                                    FundShareClass.id.in_(share_ids.values())
                                )
                            )
                        }
                        detail_url = f"{EID_DETAIL_URL}?fundId={fund_id}"
                        states, state_warnings = resolve_subscription_states(
                            announcements,
                            stored_shares,
                            subscription_date,
                            detail_url,
                        )
                        warnings.extend(
                            f"{candidate.code} {warning}"
                            for warning in state_warnings
                        )
                        synced_subscription_states = 0
                        for code, state in states.items():
                            share_id = share_ids.get(code)
                            if share_id is None:
                                continue
                            sync_subscription_state(
                                session,
                                share_id,
                                state,
                                collected_at,
                                authoritative=(
                                    not announcement_warnings
                                    and not state_warnings
                                ),
                            )
                            synced_subscription_states += 1

                    products += 1
                    shares += len(share_ids)
                    nav_rows += synced_nav
                    subscription_states += synced_subscription_states
                    return_metrics += synced_return_metrics
                except (httpx.HTTPError, RuntimeError, ValueError) as exc:
                    failures.append(f"{candidate.code} {candidate.name}: {exc}")
            if products == 0:
                session.rollback()
                reason = "; ".join(failures[:3])
                suffix = f": {reason}" if reason else ""
                raise RuntimeError(f"CSRC script F synced no products{suffix}")
            if dry_run:
                session.rollback()
            else:
                session.commit()
    return SyncStats(
        products,
        shares,
        nav_rows,
        fee_shares,
        scales,
        subscription_states,
        tuple(failures),
        tuple(warnings),
        return_metrics,
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Script E: select target off-exchange funds from the CSRC catalog"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="validate without committing"
    )
    parser.add_argument(
        "--limit", type=int, help="process only the first N candidates"
    )
    parser.add_argument(
        "--code",
        action="append",
        help="sync only this six-digit candidate code; may be repeated",
    )
    args = parser.parse_args()
    stats = run_sync(
        dry_run=args.dry_run, limit=args.limit, codes=tuple(args.code or ())
    )
    action = "validated" if args.dry_run else "synced"
    print(
        f"CSRC off-exchange script E catalog {action}: "
        f"{stats.products} products, {stats.shares} shares, "
        f"{stats.fee_shares} fee shares, {stats.scales} scales, "
        f"snapshot {stats.snapshot_date.isoformat()}, "
        f"{len(stats.failures)} failures, {len(stats.warnings)} warnings"
    )
    for failure in stats.failures:
        print(f"WARNING {failure}")
    for warning in stats.warnings:
        print(f"WARNING {warning}")


if __name__ == "__main__":
    run_tracked_sync("E", main)
